"""Frontier match + per-task hit rate + per-pair audit XLSX (P7 Pass 6 follow-up).

Per dispatcher 2026-05-08-orchestrator-cartographer-frontier-match-and-xls-mini-dispatch.md:

  Item A — Frontier match metric (3-tier reporting):
    strict           : left.primary_co == right.primary_co
    slice_primary    : left.primary_slice == right.primary_slice
    frontier         : (left.primary∪secondary slices) ∩ (right.primary∪secondary slices) ≠ ∅

  Item B — Per-task hit rate:
    For each SSDF task X (or LHS item in SCF pool) with ≥1 cross-reference resolved:
      has_hit_<tier> = any(<tier>(X, Y) for Y in X.refs)
    per_task_<tier>_rate = #{tasks with has_hit_<tier>} / #{tasks with ≥1 ref}

  Item C — XLSX (4 sheets) at
    data/p7_olir_audit/p7_v2_corrected/v7/reports/cross_validation_per_pair_audit.xlsx
"""
from __future__ import annotations
import hashlib
import json
import pathlib
import sys
import re
from collections import Counter, defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Resolvers + parsers from existing scripts
from scripts.cross_validate_ssdf_references import (
    parse_ssdf_references,
    resolve_asvs_refs,
    resolve_samm_refs,
    resolve_agile_refs,
    resolve_fpssd_refs,
    resolve_pcisslc_refs,
    resolve_nist_refs,
    SA_HUB_CONTROLS,
    NIST_PROCESS_FAMILIES,
)
from scripts.cross_validate_scf_strm_v7 import PILOT_NORMALIZERS

REPO = pathlib.Path(__file__).resolve().parents[1]
SUPPLIER = REPO / "data/p7_olir_audit/p7_v2_corrected/v7/SUPPLIER_v7_0.json"
STRM_DIR = REPO / "data/p7_olir_audit/scf_audit/scf_strm_extracted"
SSDF_TXT = pathlib.Path(
    "/Volumes/G-DRIVE/Shared/SecurityByDesign-TheoryOfEverything/"
    "ExternalSourcesInventory/sources/ssdf_sp800_218_v1_1/"
    "NIST.SP.800-218.extracted.txt"
)
ONTOLOGY_CORPUS = pathlib.Path(
    "/Volumes/G-DRIVE/Shared/SecurityByDesign-TheoryOfEverything/"
    "sbd-toe-ontology/formal/appsec_core/08-embeddings/augmented-text-corpus.json"
)
OUT_JSON = REPO / "data/p7_olir_audit/p7_v2_corrected/v7/reports/frontier_match_and_per_task_audit_v7.json"
OUT_XLSX = REPO / "data/p7_olir_audit/p7_v2_corrected/v7/reports/cross_validation_per_pair_audit.xlsx"


SLICE_LABELS = {
    "ACO-SCBI": "Supply Chain & Build Integrity",
    "ACO-IAT":  "Identity, Access & Session Trust",
    "ACO-ATB":  "Architecture & Trust Boundaries",
    "ACO-TSV":  "Testing, Security Validation",
    "ACO-TMR":  "Threat Modeling & Risk",
    "ACO-SPC":  "Secret Handling, Protected Configuration",
    "ACO-IVF":  "Input/Output Validation, Safe Parsing",
    "ACO-ITS":  "Integration Trust & Service Security",
    "ACO-RPR":  "Release, Promotion & Rollback Readiness",
    "ACO-SLG":  "Security Logging & Audit Trail",
}

# Programme-lead's specific worked-example focus
PROGRAMME_LEAD_FLAGGED = {
    ("ssdf_sp800_218_v1_1", "SSDF-PRACTICE-PW.8", "nist_sp800_53_rev5", "SP800-53-SA-11"),
}


def sha256(p: pathlib.Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def load_entity_labels() -> dict[str, str]:
    """Load short entity labels from sbd-toe-ontology embeddings corpus."""
    if not ONTOLOGY_CORPUS.exists():
        return {}
    d = json.load(ONTOLOGY_CORPUS.open())
    out = {}
    for r in d.get("records", []):
        eid = r.get("entity_id")
        if not eid:
            continue
        # First line of augmented_text after the entity-level prefix; cap at 80 chars
        txt = r.get("augmented_text", "")
        # Skip the boilerplate prefix; pick first substantive sentence
        lines = re.split(r"\.\s|\n", txt)
        label = ""
        for line in lines:
            line = line.strip()
            if line and not line.startswith("AppSec Core") and len(line) > 5:
                label = line[:80]
                break
        out[eid] = label or eid
    return out


def derive_v7_full_map(supplier: dict) -> dict[tuple[str, str], dict]:
    """For each GROUNDED v7 item, derive full primary + secondary structure with frontier slices.

    Returns: {(source, source_object_id): {
        primary_co (CO-level-preferred), primary_slice,
        primary_co_top_sim, primary_level (level of the chosen primary),
        secondary_cos (CO-level claims other than primary),
        secondary_slices (slices from any non-primary claim),
        frontier_slices (primary_slice ∪ all-claim slices),
        all_claims: [{level, target_id, slice, sim}, ...],
        source_text_excerpt, n_lifted_rows
    }}
    """
    out = {}
    for it in supplier["items"]:
        if it["final_classification"] != "GROUNDED":
            continue
        claims = it.get("claims", [])
        if not claims:
            continue
        co_level = [c for c in claims if c.get("level") == "ControlObjective"]
        primary = max(co_level, key=lambda c: c.get("similarity_score", 0.0)) if co_level \
                  else max(claims, key=lambda c: c.get("similarity_score", 0.0))
        primary_co = primary["target_id"]
        primary_slice = primary["slice"]
        primary_top_sim = float(primary.get("similarity_score", 0.0))
        primary_level = primary.get("level", "")

        secondary_cos = []
        secondary_slices = set()
        for c in claims:
            if c is primary:
                continue
            t = c["target_id"]
            if c.get("level") == "ControlObjective" and t != primary_co and t not in secondary_cos:
                secondary_cos.append(t)
            secondary_slices.add(c["slice"])

        # Frontier = union of all slices across all claims (primary + others)
        frontier_slices = {primary_slice} | {c["slice"] for c in claims}

        out[(it["source"], it["source_object_id"])] = {
            "primary_co": primary_co,
            "primary_slice": primary_slice,
            "primary_top_sim": primary_top_sim,
            "primary_level": primary_level,
            "secondary_cos": secondary_cos,
            "secondary_slices": sorted(secondary_slices - {primary_slice}),
            "frontier_slices": sorted(frontier_slices),
            "all_claims_compact": [
                {
                    "level": c.get("level"),
                    "target": c["target_id"],
                    "slice": c["slice"],
                    "sim": round(float(c.get("similarity_score", 0.0)), 3),
                }
                for c in sorted(claims, key=lambda c: -c.get("similarity_score", 0.0))[:6]
            ],
            "source_text_excerpt": (it.get("source_text") or "")[:200],
            "n_lifted_rows": len(it.get("source_lifted_rows", [])),
        }
    return out


def three_tier_match(left: dict, right: dict) -> dict:
    """Compute strict / slice_primary / frontier match for a pair (both sides must be in v7_map)."""
    l_set = set(left["frontier_slices"]) | {left["primary_slice"]}
    r_set = set(right["frontier_slices"]) | {right["primary_slice"]}
    return {
        "strict": left["primary_co"] == right["primary_co"],
        "slice_primary": left["primary_slice"] == right["primary_slice"],
        "frontier": bool(l_set & r_set),
    }


# ============================================================================
# POOL 1+2: SSDF same-level + cross-level pairs
# ============================================================================
def build_ssdf_pools(v7_map):
    """Return (same_level_pairs, cross_level_pairs).

    Each pair: dict with left, right, oracle_published_ref, ref_source, level_class, asvs_contaminated.
    Left is always SSDF; right is the referenced item.
    """
    if not SSDF_TXT.exists():
        print(f"ERR: SSDF text not found: {SSDF_TXT}", file=sys.stderr)
        sys.exit(1)
    ssdf_refs = parse_ssdf_references(SSDF_TXT)
    # Per-source resolver maps from v7_map
    src_to_map = defaultdict(dict)
    for (src, sid), v in v7_map.items():
        src_to_map[src][sid] = {
            "co": v["primary_co"],
            "slice": v["primary_slice"],
            "secondary_cos": v["secondary_cos"],
            "secondary_slices": v["secondary_slices"],
        }

    same_level_pairs = []
    cross_level_pairs = []

    def make_pair(ssdf_task, ssdf_sid, ref_source, ref, right_sid, right_pilot, level_class, asvs_contaminated=False):
        left = v7_map[("ssdf_sp800_218_v1_1", ssdf_sid)]
        right_key = (right_pilot, right_sid)
        if right_key not in v7_map:
            return None
        right = v7_map[right_key]
        return {
            "ssdf_task": ssdf_task,
            "left_pilot": "ssdf_sp800_218_v1_1",
            "left_sid": ssdf_sid,
            "right_pilot": right_pilot,
            "right_sid": right_sid,
            "left": left,
            "right": right,
            "ref_source": ref_source,
            "ref_published": ref,
            "level_class": level_class,
            "asvs_contaminated": asvs_contaminated,
        }

    # Build flat per-pilot maps for resolvers
    asvs5_map = src_to_map.get("asvs_v5_0_0", {})
    samm_map = src_to_map.get("owasp_samm_v2_1", {})
    nist_map = src_to_map.get("nist_sp800_53_rev5", {})
    pcisslc_map = src_to_map.get("pci_sslc_v1_1", {})
    fpssd_map = src_to_map.get("safecode_fpssd_2018", {})
    agile_map = src_to_map.get("safecode_agile_2012", {})

    for task_id, refs in sorted(ssdf_refs.items()):
        ssdf_sid = f"SSDF-TASK-{task_id}"
        if ("ssdf_sp800_218_v1_1", ssdf_sid) not in v7_map:
            ssdf_sid = f"SSDF-PRACTICE-{task_id}"
        if ("ssdf_sp800_218_v1_1", ssdf_sid) not in v7_map:
            # Task-to-practice rollup fallback: e.g., PW.8.1 → PW.8 (parent practice)
            # v7 substrate has only Practice-level SSDF items; refs attributed to a
            # task (PW.8.1) roll up to its parent practice (PW.8).
            parts = task_id.split(".")
            if len(parts) > 2:
                parent_practice = ".".join(parts[:2])
                ssdf_sid = f"SSDF-PRACTICE-{parent_practice}"
        if ("ssdf_sp800_218_v1_1", ssdf_sid) not in v7_map:
            continue

        # NIST 800-53
        if "SP80053" in refs:
            for r in resolve_nist_refs(refs["SP80053"], nist_map):
                # Find substrate v7 source_object_id (the resolver returns 'co' but not the original sid;
                # we must find sid via 'resolved_id' field which the resolver populates).
                rsid = r.get("resolved_id")
                if not rsid:
                    continue
                base_control = r["ref"].split("(")[0]
                if base_control in SA_HUB_CONTROLS:
                    level_class = "cross_level_hub"
                elif r["family"] in NIST_PROCESS_FAMILIES or base_control in {"SA-1", "SA-2", "SA-3", "SA-4", "SA-5", "SA-9"}:
                    level_class = "same_level_process"
                else:
                    level_class = "cross_level_impl"
                p = make_pair(task_id, ssdf_sid, "NIST-800-53", r["ref"], rsid,
                              "nist_sp800_53_rev5", level_class)
                if p:
                    if level_class == "same_level_process":
                        same_level_pairs.append(p)
                    else:
                        cross_level_pairs.append(p)

        # ASVS — always cross-level + ASVS-contaminated flag
        if "OWASPASVS" in refs:
            for r in resolve_asvs_refs(refs["OWASPASVS"], asvs5_map):
                # ASVS resolver returns 'co' but not 'resolved_id' or per-item id; find by SID lookup.
                # The resolver collects items matching prefix; we need SIDs. Re-implement here.
                pass
            # Re-resolve and capture SIDs explicitly
            for ref in (s.strip() for s in refs["OWASPASVS"].replace(';', ',').split(',')):
                if not ref or not ref[0].isdigit():
                    continue
                # Find ASVS items matching prefix
                for sid in asvs5_map:
                    for prefix in (f"ASVS-REQ-V{ref}", f"ASVS4-REQ-V{ref}"):
                        if sid == prefix or sid.startswith(prefix + "."):
                            p = make_pair(task_id, ssdf_sid, "OWASPASVS", ref, sid,
                                          "asvs_v5_0_0", "cross_level_impl",
                                          asvs_contaminated=True)
                            if p:
                                cross_level_pairs.append(p)
                            break

        # SAMM
        if "OWASPSAMM" in refs:
            for r in resolve_samm_refs(refs["OWASPSAMM"], samm_map):
                p = make_pair(task_id, ssdf_sid, "OWASPSAMM", r["ref"], r["resolved_id"],
                              "owasp_samm_v2_1", "same_level_process")
                if p:
                    same_level_pairs.append(p)

        # PCI SSLC
        if "PCISSLC" in refs:
            for r in resolve_pcisslc_refs(refs["PCISSLC"], pcisslc_map):
                p = make_pair(task_id, ssdf_sid, "PCISSLC", r["ref"], r["resolved_id"],
                              "pci_sslc_v1_1", "same_level_process")
                if p:
                    same_level_pairs.append(p)

        # SAFECode FPSSD
        if "SCFPSSD" in refs:
            for r in resolve_fpssd_refs(refs["SCFPSSD"], fpssd_map):
                p = make_pair(task_id, ssdf_sid, "SCFPSSD", r["ref"], r["resolved_id"],
                              "safecode_fpssd_2018", "same_level_process")
                if p:
                    same_level_pairs.append(p)

        # SAFECode Agile
        if "SCAGILE" in refs:
            for r in resolve_agile_refs(refs["SCAGILE"], agile_map):
                p = make_pair(task_id, ssdf_sid, "SCAGILE", r["ref"], r["resolved_id"],
                              "safecode_agile_2012", "same_level_process")
                if p:
                    same_level_pairs.append(p)

    return same_level_pairs, cross_level_pairs


# ============================================================================
# POOL 3: SCF v7 cross-pilot pairs
# ============================================================================
def build_scf_pool(v7_map):
    """Return list of cross-pilot pairs from SCF STRM extracts (Variant B)."""
    scf_inv = defaultdict(lambda: defaultdict(list))
    for f in sorted(STRM_DIR.glob("*_scf_strm_mappings.json")):
        d = json.load(f.open())
        pilot = d["pilot_id"]
        normalize = PILOT_NORMALIZERS.get(pilot)
        if not normalize:
            continue
        for m in d.get("mappings", []):
            scf_id = m.get("scf_control_id")
            if not scf_id:
                continue
            for fid in m.get("foreign_ids", []):
                for c in normalize(fid):
                    if (pilot, c) in v7_map:
                        scf_inv[scf_id][pilot].append(c)
                        break

    pairs = []
    for scf_id, p2items in scf_inv.items():
        pilots = sorted(p2items.keys())
        for i in range(len(pilots)):
            for j in range(i + 1, len(pilots)):
                seen = set()
                for ia in p2items[pilots[i]]:
                    for ib in p2items[pilots[j]]:
                        key = (pilots[i], ia, pilots[j], ib)
                        if key in seen:
                            continue
                        seen.add(key)
                        pairs.append({
                            "scf_id": scf_id,
                            "left_pilot": pilots[i],
                            "left_sid": ia,
                            "right_pilot": pilots[j],
                            "right_sid": ib,
                            "left": v7_map[(pilots[i], ia)],
                            "right": v7_map[(pilots[j], ib)],
                            "ref_source": "SCF-2026.1",
                            "ref_published": scf_id,
                            "level_class": "scf_cross_pilot",
                            "asvs_contaminated": False,
                        })
    return pairs


# ============================================================================
# Item A — pool-level 3-tier metrics
# ============================================================================
def pool_metrics(pool: list[dict], pool_label: str) -> dict:
    n = len(pool)
    if n == 0:
        return {"pool": pool_label, "n_pairs": 0}
    counts = {"strict": 0, "slice_primary": 0, "frontier": 0}
    for p in pool:
        m = three_tier_match(p["left"], p["right"])
        for k, v in m.items():
            if v:
                counts[k] += 1
    return {
        "pool": pool_label,
        "n_pairs": n,
        "n_strict": counts["strict"],
        "n_slice_primary": counts["slice_primary"],
        "n_frontier": counts["frontier"],
        "strict_pct": round(counts["strict"] / n * 100, 2),
        "slice_primary_pct": round(counts["slice_primary"] / n * 100, 2),
        "frontier_pct": round(counts["frontier"] / n * 100, 2),
    }


# ============================================================================
# Item B — per-task hit rate
# ============================================================================
def per_task_hit_rate(pool: list[dict], task_key_fn) -> dict:
    """For each task (left-side aggregator), check if ≥1 of its pairs matched at each tier."""
    by_task = defaultdict(list)
    for p in pool:
        tk = task_key_fn(p)
        by_task[tk].append(p)
    n_tasks = len(by_task)
    if n_tasks == 0:
        return {"n_tasks": 0}
    has_strict = has_slice = has_frontier = 0
    for tk, pairs in by_task.items():
        h_s = h_p = h_f = False
        for p in pairs:
            m = three_tier_match(p["left"], p["right"])
            if m["strict"]: h_s = True
            if m["slice_primary"]: h_p = True
            if m["frontier"]: h_f = True
            if h_s and h_p and h_f:
                break
        if h_s: has_strict += 1
        if h_p: has_slice += 1
        if h_f: has_frontier += 1
    return {
        "n_tasks": n_tasks,
        "n_with_strict_hit": has_strict,
        "n_with_slice_primary_hit": has_slice,
        "n_with_frontier_hit": has_frontier,
        "per_task_strict_pct": round(has_strict / n_tasks * 100, 2),
        "per_task_slice_primary_pct": round(has_slice / n_tasks * 100, 2),
        "per_task_frontier_pct": round(has_frontier / n_tasks * 100, 2),
    }


# ============================================================================
# Item C — XLSX writer
# ============================================================================
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STRICT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
SLICE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # yellow
FRONTIER_FILL = PatternFill(start_color="FFD8B5", end_color="FFD8B5", fill_type="solid") # orange
DIFF_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # red
PROG_FILL = PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid")     # purple


def match_class(m: dict) -> str:
    if m["strict"]:
        return "strict"
    elif m["slice_primary"]:
        return "slice_primary_only"
    elif m["frontier"]:
        return "frontier_only"
    else:
        return "different"


CLASS_FILL = {
    "strict": STRICT_FILL,
    "slice_primary_only": SLICE_FILL,
    "frontier_only": FRONTIER_FILL,
    "different": DIFF_FILL,
}


def write_pair_sheet(ws, pool: list[dict], entity_labels: dict[str, str], with_asvs_flag: bool, oracle_label: str):
    headers = [
        "pair_id", "oracle", "oracle_published_ref",
        "left_source", "left_id", "left_text_excerpt", "left_n_lifted_rows",
        "left_primary_CO", "left_primary_CO_label", "left_primary_slice", "left_primary_slice_label",
        "left_secondary_COs", "left_secondary_slices", "left_frontier_slices",
        "left_top_similarity", "left_primary_level",
        "right_source", "right_id", "right_text_excerpt", "right_n_lifted_rows",
        "right_primary_CO", "right_primary_CO_label", "right_primary_slice", "right_primary_slice_label",
        "right_secondary_COs", "right_secondary_slices", "right_frontier_slices",
        "right_top_similarity", "right_primary_level",
        "strict_match", "slice_match_primary", "slice_match_frontier",
        "match_class", "level_class",
    ]
    if with_asvs_flag:
        headers.append("asvs_contaminated")
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for i, p in enumerate(pool, 1):
        L, R = p["left"], p["right"]
        m = three_tier_match(L, R)
        cls = match_class(m)
        row = [
            f"{oracle_label}-{i:05d}",
            p.get("ref_source") or oracle_label,
            p.get("ref_published") or p.get("scf_id") or "",
            p["left_pilot"], p["left_sid"], L["source_text_excerpt"], L["n_lifted_rows"],
            L["primary_co"], entity_labels.get(L["primary_co"], "")[:80],
            L["primary_slice"], SLICE_LABELS.get(L["primary_slice"], ""),
            ", ".join(L["secondary_cos"]), ", ".join(L["secondary_slices"]), ", ".join(L["frontier_slices"]),
            L["primary_top_sim"], L["primary_level"],
            p["right_pilot"], p["right_sid"], R["source_text_excerpt"], R["n_lifted_rows"],
            R["primary_co"], entity_labels.get(R["primary_co"], "")[:80],
            R["primary_slice"], SLICE_LABELS.get(R["primary_slice"], ""),
            ", ".join(R["secondary_cos"]), ", ".join(R["secondary_slices"]), ", ".join(R["frontier_slices"]),
            R["primary_top_sim"], R["primary_level"],
            m["strict"], m["slice_primary"], m["frontier"],
            cls, p.get("level_class", ""),
        ]
        if with_asvs_flag:
            row.append(p.get("asvs_contaminated", False))
        ws.append(row)
        # Color match_class column (column index = 32 by default, +1 if asvs flag)
        cls_col = headers.index("match_class") + 1
        ws.cell(row=i + 1, column=cls_col).fill = CLASS_FILL.get(cls, DIFF_FILL)
        # Highlight programme-lead's PW.8 / SA-11 pair
        key = (p["left_pilot"], p["left_sid"], p["right_pilot"], p["right_sid"])
        if key in PROGRAMME_LEAD_FLAGGED or (key[2], key[3], key[0], key[1]) in PROGRAMME_LEAD_FLAGGED:
            for c in range(1, len(headers) + 1):
                ws.cell(row=i + 1, column=c).fill = PROG_FILL


def write_summary_and_examples_sheet(ws, ssdf_same, ssdf_cross, scf_pool, summary):
    ws.append(["Summary statistics — frontier match + per-task hit rate at substrate v7"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])
    # Per-pool 3-tier table
    ws.append(["", "n_pairs", "strict %", "slice_primary %", "frontier %", "n_tasks",
               "per-task strict %", "per-task slice_primary %", "per-task frontier %"])
    for col in range(1, 10):
        ws.cell(row=3, column=col).fill = HEADER_FILL
        ws.cell(row=3, column=col).font = HEADER_FONT
    pool_rows = [
        ("SSDF same-level (process→process)", "ssdf_same_level"),
        ("SSDF cross-level (incl. ASVS-contaminated subset)", "ssdf_cross_level"),
        ("SCF v7 cross-pilot pairs", "scf_v7"),
    ]
    for label, key in pool_rows:
        m = summary[key]
        t = summary[f"{key}_per_task"]
        ws.append([
            label,
            m["n_pairs"], m["strict_pct"], m["slice_primary_pct"], m["frontier_pct"],
            t["n_tasks"], t["per_task_strict_pct"], t["per_task_slice_primary_pct"], t["per_task_frontier_pct"],
        ])
    ws.append([])
    ws.append(["Diagnostic interpretation framework (per dispatcher §9-cell matrix)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Pattern observed", "Interpretation"])
    for col in range(1, 3):
        ws.cell(row=ws.max_row, column=col).fill = HEADER_FILL
        ws.cell(row=ws.max_row, column=col).font = HEADER_FONT
    ws.append(["Frontier ~60-80% + per-task hit ~70-90%",
               "Multi-claim alignment captured; primary-CO selection mathematical artifact; methodology disclosure suficiente para paper §8.2"])
    ws.append(["Frontier ~30-40% + per-task ~50%",
               "Genuíno semantic divergence — pipeline tuning OR slice taxonomy issue; investigate pre-paper claim"])
    ws.append(["Frontier ~80%+ + per-task ~95%",
               "Strong alignment confirmed; clean §8.2 story"])
    ws.append([])
    ws.append(["Worked examples — flagged pairs"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(["category", "left_pair", "right_pair", "left_primary_CO/slice", "right_primary_CO/slice",
               "left_frontier_slices", "right_frontier_slices", "match_class", "comment"])
    for col in range(1, 10):
        ws.cell(row=ws.max_row, column=col).fill = HEADER_FILL
        ws.cell(row=ws.max_row, column=col).font = HEADER_FONT

    # Pick representative examples
    examples = []
    # Programme-lead's specific pair: PW.8 / SA-11
    for p in ssdf_cross + ssdf_same:
        key = (p["left_pilot"], p["left_sid"], p["right_pilot"], p["right_sid"])
        if key in PROGRAMME_LEAD_FLAGGED or (key[2], key[3], key[0], key[1]) in PROGRAMME_LEAD_FLAGGED:
            examples.append(("⚑ programme-lead specific", p, "PW.8 / SA-11 — both expected ACO-TSV; primary-CO selection picks SCBI/TMR, frontier reaches TSV via Practice/Mechanism claims"))
            break

    def pick_n(pool, predicate, n, label, comment):
        c = 0
        for p in pool:
            m = three_tier_match(p["left"], p["right"])
            cls = match_class(m)
            if predicate(cls):
                examples.append((label, p, comment))
                c += 1
                if c >= n:
                    break

    pick_n(ssdf_same, lambda c: c == "strict", 1, "SSDF same-level strict",
           "Both sides land on identical CO — convergence at narrow band")
    pick_n(ssdf_same, lambda c: c == "slice_primary_only", 1, "SSDF same-level slice_primary",
           "Same slice, different CO — slice-bridge convergence")
    pick_n(ssdf_same, lambda c: c == "frontier_only", 2, "SSDF same-level frontier_only",
           "Multi-claim slice intersection only — primary-CO mismatched; secondary/Mechanism slices align")
    pick_n(ssdf_cross, lambda c: c == "different" and True, 1, "SSDF cross-level different",
           "Cross-level pair with no slice alignment (often asvs-contaminated)")
    pick_n(scf_pool, lambda c: c == "strict", 1, "SCF strict",
           "Cross-pilot pair landing on identical CO via SCF bridge")
    pick_n(scf_pool, lambda c: c == "frontier_only", 2, "SCF frontier_only",
           "SCF cross-pilot pair: multi-claim slice intersection")
    pick_n(scf_pool, lambda c: c == "different", 1, "SCF different",
           "SCF cross-pilot pair: SCF umbrella domain links semantically unrelated items")

    for label, p, comment in examples:
        L, R = p["left"], p["right"]
        m = three_tier_match(L, R)
        ws.append([
            label,
            f"{p['left_pilot']}/{p['left_sid']}",
            f"{p['right_pilot']}/{p['right_sid']}",
            f"{L['primary_co']} / {L['primary_slice']}",
            f"{R['primary_co']} / {R['primary_slice']}",
            ", ".join(L["frontier_slices"]),
            ", ".join(R["frontier_slices"]),
            match_class(m),
            comment,
        ])
        if "programme-lead" in label:
            for c in range(1, 10):
                ws.cell(row=ws.max_row, column=c).fill = PROG_FILL


def main():
    print("=" * 90)
    print("Frontier match + per-task hit rate + per-pair audit XLSX (P7 Pass 6 follow-up)")
    print("=" * 90)
    sup = json.load(SUPPLIER.open())
    sup_sha = sha256(SUPPLIER)
    print(f"\nSubstrate v7 SUPPLIER SHA256: {sup_sha}\n")

    print("[1/4] deriving v7 full map (primary + secondary + frontier slices)...", file=sys.stderr)
    v7_map = derive_v7_full_map(sup)
    print(f"      v7_map: {len(v7_map)} GROUNDED items", file=sys.stderr)

    print("[2/4] building SSDF pools...", file=sys.stderr)
    ssdf_same, ssdf_cross = build_ssdf_pools(v7_map)
    print(f"      SSDF same-level: {len(ssdf_same)} pairs", file=sys.stderr)
    print(f"      SSDF cross-level: {len(ssdf_cross)} pairs ({sum(1 for p in ssdf_cross if p['asvs_contaminated'])} ASVS-contaminated)", file=sys.stderr)

    print("[3/4] building SCF v7 pool...", file=sys.stderr)
    scf_pool = build_scf_pool(v7_map)
    print(f"      SCF v7: {len(scf_pool)} cross-pilot pairs", file=sys.stderr)

    # Item A: pool-level 3-tier metrics
    summary = {
        "ssdf_same_level": pool_metrics(ssdf_same, "ssdf_same_level"),
        "ssdf_cross_level": pool_metrics(ssdf_cross, "ssdf_cross_level"),
        "scf_v7": pool_metrics(scf_pool, "scf_v7"),
    }

    # Item A bonus: ASVS-contaminated subset of cross-level (separately reported)
    asvs_only = [p for p in ssdf_cross if p["asvs_contaminated"]]
    non_asvs_cross = [p for p in ssdf_cross if not p["asvs_contaminated"]]
    summary["ssdf_cross_level_asvs_contaminated"] = pool_metrics(asvs_only, "ssdf_cross_level_asvs_contaminated")
    summary["ssdf_cross_level_non_asvs"] = pool_metrics(non_asvs_cross, "ssdf_cross_level_non_asvs")

    # Item B: per-task hit rates
    summary["ssdf_same_level_per_task"] = per_task_hit_rate(ssdf_same, lambda p: (p["left_pilot"], p["left_sid"]))
    summary["ssdf_cross_level_per_task"] = per_task_hit_rate(ssdf_cross, lambda p: (p["left_pilot"], p["left_sid"]))
    summary["scf_v7_per_task"] = per_task_hit_rate(scf_pool, lambda p: (p["left_pilot"], p["left_sid"]))

    print()
    print("Item A — 3-tier per-pair metrics:")
    print(f"{'pool':<55}  {'n':>5}  {'strict':>7}  {'slice':>7}  {'frontier':>9}")
    for k in ["ssdf_same_level", "ssdf_cross_level", "ssdf_cross_level_asvs_contaminated",
              "ssdf_cross_level_non_asvs", "scf_v7"]:
        m = summary[k]
        print(f"  {k:<53}  {m['n_pairs']:>5}  {m.get('strict_pct', 0):>5.2f}%  {m.get('slice_primary_pct', 0):>5.2f}%  {m.get('frontier_pct', 0):>7.2f}%")

    print()
    print("Item B — per-task hit rates:")
    print(f"{'pool':<55}  {'n_tasks':>7}  {'strict':>7}  {'slice':>7}  {'frontier':>9}")
    for k in ["ssdf_same_level", "ssdf_cross_level", "scf_v7"]:
        t = summary[f"{k}_per_task"]
        print(f"  {k:<53}  {t['n_tasks']:>7}  {t['per_task_strict_pct']:>5.2f}%  {t['per_task_slice_primary_pct']:>5.2f}%  {t['per_task_frontier_pct']:>7.2f}%")

    # Confirm PW.8 / SA-11 in cross_level (or same_level)
    pwsa = [p for p in ssdf_cross + ssdf_same
            if p["left_sid"] == "SSDF-PRACTICE-PW.8" and p["right_sid"] == "SP800-53-SA-11"]
    if pwsa:
        p = pwsa[0]
        m = three_tier_match(p["left"], p["right"])
        print()
        print(f"⚑ PW.8 / SA-11 pair found in pool '{p['level_class']}'")
        print(f"   left  primary={p['left']['primary_co']} slice={p['left']['primary_slice']} frontier={p['left']['frontier_slices']}")
        print(f"   right primary={p['right']['primary_co']} slice={p['right']['primary_slice']} frontier={p['right']['frontier_slices']}")
        print(f"   strict={m['strict']} slice_primary={m['slice_primary']} frontier={m['frontier']}")

    # Item C: XLSX
    print("\n[4/4] writing XLSX...", file=sys.stderr)
    entity_labels = load_entity_labels()
    wb = Workbook()
    # Sheet 1
    ws1 = wb.active
    ws1.title = "1_SSDF_same_level"
    write_pair_sheet(ws1, ssdf_same, entity_labels, with_asvs_flag=False, oracle_label="SSDF-SL")
    # Sheet 2
    ws2 = wb.create_sheet("2_SSDF_cross_level")
    write_pair_sheet(ws2, ssdf_cross, entity_labels, with_asvs_flag=True, oracle_label="SSDF-CL")
    # Sheet 3
    ws3 = wb.create_sheet("3_SCF_v7")
    write_pair_sheet(ws3, scf_pool, entity_labels, with_asvs_flag=False, oracle_label="SCF")
    # Sheet 4 — summary
    ws4 = wb.create_sheet("4_summary_and_examples")
    write_summary_and_examples_sheet(ws4, ssdf_same, ssdf_cross, scf_pool, summary)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"[write] {OUT_XLSX.relative_to(REPO)}")

    # Item A+B JSON
    out = {
        "schema": "frontier_match_and_per_task_audit_v7/1.0",
        "substrate_supplier_sha256": sup_sha,
        "summary": summary,
        "ssdf_pw8_sa11_pair_check": (
            None if not pwsa else {
                "found_in_pool": pwsa[0]["level_class"],
                "left_primary_co": pwsa[0]["left"]["primary_co"],
                "left_primary_slice": pwsa[0]["left"]["primary_slice"],
                "left_frontier_slices": pwsa[0]["left"]["frontier_slices"],
                "right_primary_co": pwsa[0]["right"]["primary_co"],
                "right_primary_slice": pwsa[0]["right"]["primary_slice"],
                "right_frontier_slices": pwsa[0]["right"]["frontier_slices"],
                "strict": three_tier_match(pwsa[0]["left"], pwsa[0]["right"])["strict"],
                "slice_primary": three_tier_match(pwsa[0]["left"], pwsa[0]["right"])["slice_primary"],
                "frontier": three_tier_match(pwsa[0]["left"], pwsa[0]["right"])["frontier"],
            }
        ),
    }
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"[write] {OUT_JSON.relative_to(REPO)}")


if __name__ == "__main__":
    main()
